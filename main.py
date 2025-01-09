import os
import logging
import schedule
import time
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
from telegram.ext import ConversationHandler, CallbackQueryHandler
from openpyxl import Workbook, load_workbook

# Setting up logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

# Excel file for storing appointments
EXCEL_FILE = 'appointments.xlsx'

# Enter the Telegram bot token
TELEGRAM_TOKEN = 'YOUR TELEGRAM TOKEN'

# Global variables for conversation states
ID, FIRST_NAME, DATE, TIME, REASON, PHONE_NUMBER, CONFIRMATION = range(7)

# Ensure the Excel file exists with the correct structure
def initialize_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Appointments"
        sheet.append(["ID", "First Name", "Date", "Time", "Reason", "Phone Number"])  # Header row
        workbook.save(EXCEL_FILE)

# Write a new appointment to the Excel file
def write_appointment_to_excel(data):
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook["Appointments"]
    sheet.append(data)
    workbook.save(EXCEL_FILE)

# Command to start the bot conversation
async def start(update: Update, context: CallbackContext):
    await update.message.reply_text("Welcome to the Appointment Booking Bot! Please provide your ID.")
    return ID

# Handle the user's ID input
async def id_input(update: Update, context: CallbackContext):
    user_id = update.message.text
    context.user_data['id'] = user_id
    await update.message.reply_text(f"Thanks! Now, please provide your first name.")
    return FIRST_NAME

# Handle the user's First Name input
async def first_name_input(update: Update, context: CallbackContext):
    first_name = update.message.text
    context.user_data['first_name'] = first_name
    await update.message.reply_text(
        f"Got it, {first_name}! Now, please select a date for your appointment (e.g., 2025-01-05).")
    return DATE

# Handle the appointment date input
async def date_input(update: Update, context: CallbackContext):
    appointment_date = update.message.text
    context.user_data['date'] = appointment_date
    await update.message.reply_text(f"Great! Now, please choose a time (e.g., 14:30).")
    return TIME

# Handle the appointment time input
async def time_input(update: Update, context: CallbackContext):
    appointment_time = update.message.text
    context.user_data['time'] = appointment_time
    await update.message.reply_text("Please tell me the reason for the appointment.")
    return REASON

# Handle the reason input
async def reason_input(update: Update, context: CallbackContext):
    reason = update.message.text
    context.user_data['reason'] = reason
    await update.message.reply_text("Please provide your phone number (e.g., 0712345678).")
    return PHONE_NUMBER

# Handle the phone number input
async def phone_number_input(update: Update, context: CallbackContext):
    phone_number = update.message.text
    context.user_data['phone_number'] = phone_number

    # Create inline buttons for "Confirm" and "Cancel"
    keyboard = [
        [
            InlineKeyboardButton("Confirm", callback_data="confirm"),
            InlineKeyboardButton("Cancel", callback_data="cancel")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(f"Please confirm your appointment details:\n"
                                    f"ID: {context.user_data['id']}\n"
                                    f"Name: {context.user_data['first_name']}\n"
                                    f"Date: {context.user_data['date']}\n"
                                    f"Time: {context.user_data['time']}\n"
                                    f"Reason: {context.user_data['reason']}\n"
                                    f"Phone Number: {context.user_data['phone_number']}\n",
                                    reply_markup=reply_markup)
    return CONFIRMATION

# Handle confirmation via buttons
async def button_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    # Get the callback data ("confirm" or "cancel")
    choice = query.data

    if choice == "confirm":
        # Write appointment data to Excel
        data = [
            context.user_data['id'],
            context.user_data['first_name'],
            context.user_data['date'],
            context.user_data['time'],
            context.user_data['reason'],
            context.user_data['phone_number']
        ]
        write_appointment_to_excel(data)  # Save appointment to the Excel file

        await query.edit_message_text("Your appointment has been confirmed!")

        # Set up reminder notifications
        schedule_reminder(query, context.user_data['date'], context.user_data['time'])

        return ConversationHandler.END
    elif choice == "cancel":
        await query.edit_message_text("Appointment canceled.")
        return ConversationHandler.END

# Send reminder notifications
def send_reminder(update, date, time, hours_before):
    appointment_datetime = datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M")
    reminder_time = appointment_datetime - timedelta(hours=hours_before)

    # Schedule reminder
    def reminder_job():
        update.message.reply_text(f"Reminder: Your appointment is in {hours_before} hours!")

    schedule.every().day.at(reminder_time.strftime("%H:%M")).do(reminder_job)

# Schedule reminders
def schedule_reminder(update, date, time):
    # Send reminders at 12, 3, and 1 hour before the appointment
    send_reminder(update, date, time, 12)
    send_reminder(update, date, time, 3)
    send_reminder(update, date, time, 1)

# Cancel the conversation
async def cancel(update: Update, context: CallbackContext):
    await update.message.reply_text("Appointment booking canceled.")
    return ConversationHandler.END

# Main function to run the bot
def main():
    initialize_excel_file()  # Ensure the Excel file exists before starting the bot

    application = Application.builder().token(TELEGRAM_TOKEN).build()

    # Define conversation handler
    conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, id_input)],
            FIRST_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, first_name_input)],
            DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, date_input)],
            TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, time_input)],
            REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, reason_input)],
            PHONE_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, phone_number_input)],
            CONFIRMATION: [CallbackQueryHandler(button_handler)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    application.add_handler(conversation_handler)

    # Start the bot
    application.run_polling()

if __name__ == '__main__':
    main()
